#!/usr/bin/env python

from openpyxl import Workbook
import os
import sys
wb = Workbook()
path = sys.argv[1]
name = sys.argv[2]
#print(os.listdir())
lis = os.listdir(path)
sheetnumber = 1
for x in lis:
    if os.path.isfile(x)==True:
        extension = os.path.splitext(x)
        print(extension[1])
        if extension[1] == '.txt':
            #print("Yes")
            ws = wb.create_sheet(index=0,title=extension[0])
            xyz = 1
            pqr = 1
            a = open(x)
            
            while True:
                a1 = a.readline()
                if len(a1)==0:
                    break
                data = a1.split("\t")
                for z in data:
                    r = unicode(z, 'utf-8')
                    if r.isdigit() == False: #text
                        ws.cell(row = xyz,column = pqr).value = r
                        pqr += 1
                    else:    
                        y = float(r)
                        ws.cell(row = xyz,column = pqr).value = y                             
                        pqr += 1
                pqr = 1
                xyz += 1
            sheetnumber+=1
    else:
        pass
wb.save(name+'.xlsx')

